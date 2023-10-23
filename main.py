import os
import platform
import re
import subprocess
from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def browse_files():
    """Opening the file explorer window."""
    filename = filedialog.askopenfilename(initialdir="/",
                                          title="Select a File",
                                          filetypes=(("Spreadsheets", "*.xlsx *.xlsm *.xltx *.xltm"),
                                                     ("All files", "*.*")))
    # Change label contents
    label_file_explorer.configure(text=filename)


def open_output_txt_file():
    """Open the links.txt output file"""
    try:
        if platform.system() == 'Darwin':  # macOS
            subprocess.call(('open', "links.txt"))
        elif platform.system() == 'Windows':  # Windows
            os.startfile("links.txt")
        else:  # linux variants
            subprocess.call(('xdg-open', "links.txt"))
    except FileNotFoundError:
        output.configure(text=f"Can't find the links.txt output file.\n")


def find_links(file: str) -> None:
    """Extracting URLs from the file"""
    # Open the file.
    try:
        wb = load_workbook(file, data_only=True)
    except InvalidFileException:
        return output.configure(text=f'Please select a valid file.\n')

    # Go through all the sheets and cells of the file.
    links = []
    regex = r'(?i)\b((?:https?:(?:/{1,3}|[a-z0-9%])|[a-z0-9.\-]+[.](?:com|net|org|edu|gov|mil|aero|asia|biz|cat|coop|' \
            r'info|int|jobs|mobi|museum|name|post|pro|tel|travel|xxx|ac|ad|ae|af|ag|ai|al|am|an|ao|aq|ar|as|at|au|aw|' \
            r'ax|az|ba|bb|bd|be|bf|bg|bh|bi|bj|bm|bn|bo|br|bs|bt|bv|bw|by|bz|ca|cc|cd|cf|cg|ch|ci|ck|cl|cm|cn|co|cr|cs' \
            r'|cu|cv|cx|cy|cz|dd|de|dj|dk|dm|do|dz|ec|ee|eg|eh|er|es|et|eu|fi|fj|fk|fm|fo|fr|ga|gb|gd|ge|gf|gg|gh|gi|gl' \
            r'|gm|gn|gp|gq|gr|gs|gt|gu|gw|gy|hk|hm|hn|hr|ht|hu|id|ie|il|im|in|io|iq|ir|is|it|je|jm|jo|jp|ke|kg|kh|ki|km' \
            r'|kn|kp|kr|kw|ky|kz|la|lb|lc|li|lk|lr|ls|lt|lu|lv|ly|ma|mc|md|me|mg|mh|mk|ml|mm|mn|mo|mp|mq|mr|ms|mt|mu|mv' \
            r'|mw|mx|my|mz|na|nc|ne|nf|ng|ni|nl|no|np|nr|nu|nz|om|pa|pe|pf|pg|ph|pk|pl|pm|pn|pr|ps|pt|pw|py|qa|re|ro|rs' \
            r'|ru|rw|sa|sb|sc|sd|se|sg|sh|si|sj|Ja|sk|sl|sm|sn|so|sr|ss|st|su|sv|sx|sy|sz|tc|td|tf|tg|th|tj|tk|tl|tm|tn|' \
            r'to|tp|tr|tt|tv|tw|tz|ua|ug|uk|us|uy|uz|va|vc|ve|vg|vi|vn|vu|wf|ws|ye|yt|yu|za|zm|zw)/)(?:[^\s()<>{}\[\]]+|' \
            r'\([^\s()]*?\([^\s()]+\)[^\s()]*?\)|\([^\s]+?\))+(?:\([^\s()]*?\([^\s()]+\)[^\s()]*?\)|\([^\s]+?\)|[^\s`!()' \
            r'\[\]{};:\'".,<>?«»“”‘’])|(?:(?<!@)[a-z0-9]+(?:[.\-][a-z0-9]+)*[.](?:com|net|org|edu|gov|mil|aero|asia|biz|' \
            r'cat|coop|info|int|jobs|mobi|museum|name|post|pro|tel|travel|xxx|ac|ad|ae|af|ag|ai|al|am|an|ao|aq|ar|as|at|' \
            r'au|aw|ax|az|ba|bb|bd|be|bf|bg|bh|bi|bj|bm|bn|bo|br|bs|bt|bv|bw|by|bz|ca|cc|cd|cf|cg|ch|ci|ck|cl|cm|cn|co|' \
            r'cr|cs|cu|cv|cx|cy|cz|dd|de|dj|dk|dm|do|dz|ec|ee|eg|eh|er|es|et|eu|fi|fj|fk|fm|fo|fr|ga|gb|gd|ge|gf|gg|gh|' \
            r'gi|gl|gm|gn|gp|gq|gr|gs|gt|gu|gw|gy|hk|hm|hn|hr|ht|hu|id|ie|il|im|in|io|iq|ir|is|it|je|jm|jo|jp|ke|kg|kh|' \
            r'ki|km|kn|kp|kr|kw|ky|kz|la|lb|lc|li|lk|lr|ls|lt|lu|lv|ly|ma|mc|md|me|mg|mh|mk|ml|mm|mn|mo|mp|mq|mr|ms|mt|' \
            r'mu|mv|mw|mx|my|mz|na|nc|ne|nf|ng|ni|nl|no|np|nr|nu|nz|om|pa|pe|pf|pg|ph|pk|pl|pm|pn|pr|ps|pt|pw|py|qa|re|' \
            r'ro|rs|ru|rw|sa|sb|sc|sd|se|sg|sh|si|sj|Ja|sk|sl|sm|sn|so|sr|ss|st|su|sv|sx|sy|sz|tc|td|tf|tg|th|tj|tk|tl|' \
            r'tm|tn|to|tp|tr|tt|tv|tw|tz|ua|ug|uk|us|uy|uz|va|vc|ve|vg|vi|vn|vu|wf|ws|ye|yt|yu|za|zm|zw)\b/?(?!@)))'
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:  # Filter the value in every cell.
                    url = re.findall(regex, str(cell.value))
                    url = str(url).strip("['']")
                    # Check if link is not already added and if the color fill is not yellow.
                    if url not in links and cell.fill.start_color.index != 'FFFFFF00':
                        links.append(url)

    # Go through the elements in the list and check if element is not empty. Then export them to links.txt.
    count = 0
    with open('links.txt', 'w') as f:
        for link in sorted(links):
            if link:
                count += 1
                f.write(f'{link}\n')
    # Change label contents
    output.configure(text=f'{count} unique links have been exported to the links.txt file.')


# Function for closing the window
def close_window():
    window.destroy()


# Defining Tkinter parameters
window = Tk()
window.title('URL extract from spreadsheet')
window.geometry("600x200")
window.config(background="white")

# Create a File Explorer label
label_file_explorer = Label(window, text="Please select a spreadsheet file to extract URLs from.",
                            fg="black", wraplength=585)

# Create buttons and output label
button_explore = Button(window, text="Browse Files", command=browse_files)
button_extract = Button(window, text="Extract", command=lambda: find_links(label_file_explorer.cget("text")))
open_txt_file = Button(window, text="Open links.txt", command=lambda: open_output_txt_file())
output = Label(window, height=5, width=74, fg="black", wraplength=585)
button_exit = Button(window, text="Exit", command=close_window)

# Place the elements in the window
label_file_explorer.grid(column=0, row=0, sticky=EW, columnspan=3)
button_explore.grid(column=0, row=1, sticky=E, padx=0, pady=5)
button_extract.grid(column=1, row=1, sticky=EW, padx=0, pady=5)
open_txt_file.grid(column=2, row=1, sticky=W, padx=0, pady=5)
output.grid(column=0, row=2, sticky=W, columnspan=3)
button_exit.grid(column=0, row=3, sticky=EW, columnspan=3)

# Let the window wait for any events
window.mainloop()
